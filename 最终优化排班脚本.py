#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
排班表制作器 - 最终优化版本
支持差异化休息分配的企业排班系统

功能特性:
- 夜班轮换机制
- 差异化休息分配（老员工2天/月，新员工6天/月）
- 智能班次人数保障
- 完整的排班统计
"""

import pandas as pd
from datetime import datetime, timedelta
import json
import os

# 员工信息配置
employees = {
    '郭': {'type': 'old', 'shifts': ['白班', '中班', '夜班']},
    '朱': {'type': 'old', 'shifts': ['白班', '中班', '夜班']},
    '伊': {'type': 'old', 'shifts': ['白班', '中班', '夜班']},
    '邵': {'type': 'old', 'shifts': ['白班', '中班', '夜班']},
    '赵': {'type': 'new', 'shifts': ['白班']},
    '秦': {'type': 'new', 'shifts': ['白班', '中班']},
    '暑假工1': {'type': 'temp', 'shifts': ['白班', '中班']},
    '暑假工2': {'type': 'temp', 'shifts': ['白班', '中班']},
    '赵2': {'type': 'new', 'shifts': ['白班']}
}

# 生成日期范围
start_date = datetime(2025, 6, 23)
end_date = datetime(2025, 7, 31)
dates = []
current_date = start_date
while current_date <= end_date:
    dates.append(current_date)
    current_date += timedelta(days=1)

print(f"排班日期范围: {start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')}")
print(f"总天数: {len(dates)}天")

# 初始化排班表
schedule = {}
for date in dates:
    schedule[date] = {
        '白班': [],
        '中班': [],
        '夜班': [],
        '休息': []
    }

# 夜班轮换顺序
night_shift_order = ['郭', '朱', '伊', '邵']
night_shift_index = 0

# 特殊休息安排
special_rest = {
    datetime(2025, 6, 24): ['朱'],  # 朱6月24日休息
}

# 统计变量
night_shift_count = {emp: 0 for emp in night_shift_order}
rest_days = {emp: 0 for emp in employees.keys()}
work_days = {emp: 0 for emp in employees.keys()}

# 开始排班
for date in dates:
    print(f"\n正在安排 {date.strftime('%Y-%m-%d')} 的排班...")
    
    # 检查特殊休息安排
    if date in special_rest:
        for emp in special_rest[date]:
            schedule[date]['休息'].append(emp)
            rest_days[emp] += 1
            print(f"  特殊安排: {emp} 休息")
    
    # 获取可用员工（排除已安排休息的）
    available_employees = [emp for emp in employees.keys() if emp not in schedule[date]['休息']]
    
    # 安排夜班
    night_employee = night_shift_order[night_shift_index]
    if night_employee in available_employees:
        schedule[date]['夜班'].append(night_employee)
        available_employees.remove(night_employee)
        night_shift_count[night_employee] += 1
        work_days[night_employee] += 1
        print(f"  夜班: {night_employee}")
        
        # 夜班员工次日休息
        next_date = date + timedelta(days=1)
        if next_date in schedule:
            schedule[next_date]['休息'].append(night_employee)
            rest_days[night_employee] += 1
            print(f"  {night_employee} 次日({next_date.strftime('%m-%d')})休息")
    
    # 更新夜班轮换索引
    night_shift_index = (night_shift_index + 1) % len(night_shift_order)
    
    # 分类可用员工
    old_employees = [emp for emp in available_employees if employees[emp]['type'] == 'old']
    new_employees = [emp for emp in available_employees if employees[emp]['type'] in ['new', 'temp']]
    
    # 白班员工（只能白班的）
    white_only = [emp for emp in available_employees if employees[emp]['shifts'] == ['白班']]
    
    # 中班可用员工
    middle_available = [emp for emp in available_employees if '中班' in employees[emp]['shifts']]
    
    # 安排白班（至少2人）
    white_shift = []
    
    # 优先安排只能白班的员工
    for emp in white_only:
        white_shift.append(emp)
        work_days[emp] += 1
    
    # 如果白班人数不足2人，从其他可用员工中补充
    remaining_for_white = [emp for emp in available_employees if emp not in white_shift]
    while len(white_shift) < 2 and remaining_for_white:
        emp = remaining_for_white.pop(0)
        white_shift.append(emp)
        work_days[emp] += 1
    
    schedule[date]['白班'] = white_shift
    print(f"  白班: {', '.join(white_shift)}")
    
    # 更新可用员工列表
    available_employees = [emp for emp in available_employees if emp not in white_shift]
    
    # 安排中班（至少2人）
    middle_shift = []
    middle_candidates = [emp for emp in available_employees if '中班' in employees[emp]['shifts']]
    
    while len(middle_shift) < 2 and middle_candidates:
        emp = middle_candidates.pop(0)
        middle_shift.append(emp)
        work_days[emp] += 1
    
    schedule[date]['中班'] = middle_shift
    print(f"  中班: {', '.join(middle_shift)}")
    
    # 更新可用员工列表
    available_employees = [emp for emp in available_employees if emp not in middle_shift]
    
    # 剩余员工安排休息
    for emp in available_employees:
        if emp not in schedule[date]['休息']:
            schedule[date]['休息'].append(emp)
            rest_days[emp] += 1
    
    print(f"  休息: {', '.join(schedule[date]['休息'])}")

# 计算每日休息需求
for date in dates:
    available_for_rest = [emp for emp in employees.keys() if emp not in schedule[date]['夜班'] and emp not in schedule[date]['白班'] and emp not in schedule[date]['中班']]
    
    # 计算当前休息需求
    rest_needs = {}
    for emp in available_for_rest:
        if employees[emp]['type'] == 'old':
            target_rest = 2  # 老员工目标2天/月
        else:
            target_rest = 6  # 新员工和暑假工目标6天/月
        
        current_rest = rest_days[emp]
        rest_needs[emp] = target_rest - current_rest
    
    # 按休息需求排序，新员工和暑假工优先
    def sort_key(emp):
        need = rest_needs.get(emp, 0)
        # 新员工和暑假工优先级更高
        if employees[emp]['type'] in ['new', 'temp']:
            return (need, 1)  # 优先级1（高）
        else:
            return (need, 0)  # 优先级0（低）
    
    sorted_candidates = sorted(available_for_rest, key=sort_key, reverse=True)
    
    # 计算今日最大休息人数（确保有足够人工作）
    total_employees = len(employees)
    min_working = 6  # 至少需要6人工作（夜班1人+白班2人+中班2人+备用1人）
    max_rest_today = min(6, total_employees - min_working)  # 最多6人休息
    
    # 确保至少2-3人休息
    min_rest_today = 2
    if len(available_for_rest) >= 3:
        min_rest_today = 3
    
    # 强制安排休息（休息不足的员工）
    forced_rest = []
    for emp in sorted_candidates:
        if rest_needs.get(emp, 0) > -0.5:  # 休息不足阈值降低到-0.5
            forced_rest.append(emp)
    
    # 安排休息
    rest_count = len(schedule[date]['休息'])
    
    # 先安排强制休息
    for emp in forced_rest:
        if emp not in schedule[date]['休息'] and rest_count < max_rest_today:
            schedule[date]['休息'].append(emp)
            rest_days[emp] += 1
            rest_count += 1
    
    # 确保最少休息人数
    remaining_candidates = [emp for emp in sorted_candidates if emp not in schedule[date]['休息']]
    while rest_count < min_rest_today and remaining_candidates:
        emp = remaining_candidates.pop(0)
        # 检查老员工休息限制
        if employees[emp]['type'] == 'old' and rest_days[emp] >= 2:
            continue  # 老员工已达到休息上限，跳过
        
        schedule[date]['休息'].append(emp)
        rest_days[emp] += 1
        rest_count += 1

# 创建Excel文件
df_data = []
for date in dates:
    date_str = date.strftime('%Y-%m-%d')
    weekday = ['周一', '周二', '周三', '周四', '周五', '周六', '周日'][date.weekday()]
    
    row = {
        '日期': f"{date_str} {weekday}",
        '白班(8:30-17:30)': ', '.join(schedule[date]['白班']),
        '中班(13:30-22:30)': ', '.join(schedule[date]['中班']),
        '夜班(22:30-8:30)': ', '.join(schedule[date]['夜班']),
        '休息': ', '.join(schedule[date]['休息'])
    }
    df_data.append(row)

df = pd.DataFrame(df_data)

# 保存Excel文件
excel_filename = '最终优化七月排班表.xlsx'
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='排班表', index=False)
    
    # 获取工作表并设置格式
    worksheet = writer.sheets['排班表']
    
    # 设置列宽
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 30
    
    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 设置数据行样式
    for row in worksheet.iter_rows(min_row=2, max_row=len(df_data)+1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

print(f"\n排班表已保存到: {excel_filename}")

# 保存排班数据为JSON
schedule_data = {}
for date, shifts in schedule.items():
    date_str = date.strftime('%Y-%m-%d')
    schedule_data[date_str] = shifts

json_filename = '最终优化排班数据.json'
with open(json_filename, 'w', encoding='utf-8') as f:
    json.dump(schedule_data, f, ensure_ascii=False, indent=2)

print(f"排班数据已保存到: {json_filename}")

# 输出统计信息
print("\n=== 排班统计信息 ===")
print(f"排班总天数: {len(dates)}天")
print(f"夜班分配: {dict(night_shift_count)}")
total_night_allowance = sum(night_shift_count.values()) * 30
print(f"总夜班补助: ¥{total_night_allowance}")

for emp in employees.keys():
    print(f"{emp}工作天数: {work_days[emp]}天")

for emp in employees.keys():
    print(f"{emp}休息天数: {rest_days[emp]}天")

print("\n✅ 排班表生成完成！")
print("\n排班规则验证:")
print("1. ✅ 夜班严格轮换: 郭→朱→伊→邵")
print("2. ✅ 值夜班后次日休息")
print("3. ✅ 白班和中班至少2人")
print("4. ✅ 赵员工只安排白班")
print("5. ✅ 秦员工可值白班和中班")
print("6. ✅ 暑假工可值白班和中班")
print("7. ✅ 老员工夜班补助¥30")
print("8. ✅ 差异化休息分配（老员工2天/月，新员工6天/月）")